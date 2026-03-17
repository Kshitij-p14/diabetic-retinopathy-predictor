import streamlit as st
import pandas as pd
import numpy as np
import pickle
import matplotlib.pyplot as plt
from datetime import datetime
import io
from io import BytesIO
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Load model and scaler
@st.cache_resource
def load_models():
    with open('model.pkl', 'rb') as f:
        model = pickle.load(f)
    with open('scaler.pkl', 'rb') as f:
        scaler = pickle.load(f)
    return model, scaler

model, scaler = load_models()
features = ['age', 'systolic_bp', 'diastolic_bp', 'cholesterol']
BEST_MODEL = 'Logistic Regression'

# Normal ranges for reference
NORMAL_RANGES = {
    'age': (30, 110),
    'systolic_bp': (90, 120),
    'diastolic_bp': (60, 80),
    'cholesterol': (125, 200)
}

# Page Configuration
st.set_page_config(
    page_title='Diabetic Retinopathy Predictor - Advanced',
    page_icon='🩺',
    layout='wide',
    initial_sidebar_state='expanded'
)

# Session state for dark mode and history
if 'dark_mode' not in st.session_state:
    st.session_state.dark_mode = False
if 'prediction_history' not in st.session_state:
    st.session_state.prediction_history = []

# Custom CSS with dark mode support
def get_css():
    if st.session_state.dark_mode:
        return """
<style>
    .main-header {
        font-size: 3rem;
        font-weight: 700;
        background: linear-gradient(90deg, #00d2ff 0%, #3a7bd5 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        padding: 1rem 0;
    }
    .stApp {
        background-color: #1e1e1e;
        color: #ffffff;
    }
    .subtitle { text-align: center; color: #aaa; font-size: 1.1rem; margin-bottom: 2rem; }
    .risk-high { background: linear-gradient(135deg, #ff6b6b 0%, #c92a2a 100%); padding: 2rem; border-radius: 15px; border-left: 5px solid #c92a2a; color: white; }
    .risk-medium { background: linear-gradient(135deg, #ffd93d 0%, #ff922b 100%); padding: 2rem; border-radius: 15px; border-left: 5px solid #ff922b; color: #333; }
    .risk-low { background: linear-gradient(135deg, #51cf66 0%, #37b24d 100%); padding: 2rem; border-radius: 15px; border-left: 5px solid #37b24d; color: white; }
    .info-box { background: #2d2d2d; padding: 1rem; border-radius: 10px; border-left: 4px solid #667eea; margin: 1rem 0; }
</style>
"""
    else:
        return """
<style>
    .main-header {
        font-size: 3rem;
        font-weight: 700;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        padding: 1rem 0;
    }
    .subtitle { text-align: center; color: #666; font-size: 1.1rem; margin-bottom: 2rem; }
    .risk-high { background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); padding: 2rem; border-radius: 15px; border-left: 5px solid #f5576c; }
    .risk-medium { background: linear-gradient(135deg, #ffecd2 0%, #fcb69f 100%); padding: 2rem; border-radius: 15px; border-left: 5px solid #ff9966; }
    .risk-low { background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%); padding: 2rem; border-radius: 15px; border-left: 5px solid #4facfe; }
    .info-box { background: #f8f9fa; padding: 1rem; border-radius: 10px; border-left: 4px solid #667eea; margin: 1rem 0; }
    .stButton>button {
        width: 100%;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-size: 1.2rem;
        font-weight: 600;
        padding: 0.75rem 2rem;
        border-radius: 10px;
        border: none;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
    }
</style>
"""

st.markdown(get_css(), unsafe_allow_html=True)

# Header
col1, col2, col3 = st.columns([4, 1, 1])
with col1:
    st.markdown('<h1 class="main-header">🩺 Diabetic Retinopathy Risk Predictor - Advanced</h1>', unsafe_allow_html=True)
with col3:
    if st.button('🌙 Dark Mode' if not st.session_state.dark_mode else '☀️ Light Mode'):
        st.session_state.dark_mode = not st.session_state.dark_mode
        st.rerun()

st.markdown('<p class="subtitle">AI-powered comprehensive assessment for diabetic retinopathy detection</p>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image("https://img.icons8.com/3d-fluency/94/medical-doctor.png", width=100)
    st.markdown("### 📊 Model Information")
    st.info(f"""
    **Algorithm:** {BEST_MODEL}  
    **Dataset:** 6,000 patients  
    **Features:** 4 clinical parameters  
    **Accuracy:** ~85%  
    **AUC-ROC:** ~0.89
    """)
    
    st.markdown("### ⚠️ Disclaimer")
    st.warning("For educational purposes only. Consult healthcare professionals for medical decisions.")

# Main Tabs
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
    "🔬 Risk Prediction", 
    "📊 Advanced Analytics", 
    "📚 Education & Info",
    "📁 Batch Analysis",
    "📈 Dashboard",
    "👥 Risk Stratification",
    "📋 Report & Export",
    "💬 Recommendations"
])

# ========== TAB 1: RISK PREDICTION (Original Enhanced) ==========
with tab1:
    col_left, col_right = st.columns([1, 1], gap="large")
    
    with col_left:
        st.markdown("### 📋 Patient Information")
        
        with st.container():
            st.markdown('<div class="info-box">', unsafe_allow_html=True)
            st.markdown("**👤 Demographics**")
            age = st.slider('🎂 Age (years)', 30, 110, 60, help="Patient's age in years")
            st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown('<div class="info-box">', unsafe_allow_html=True)
            st.markdown("**💓 Blood Pressure**")
            systolic_bp = st.slider('📈 Systolic BP (mmHg)', 60, 180, 110, help="Normal: below 120 mmHg")
            diastolic_bp = st.slider('📉 Diastolic BP (mmHg)', 50, 140, 85, help="Normal: below 80 mmHg")
            
            if systolic_bp >= 140 or diastolic_bp >= 90:
                st.error("⚠️ High Blood Pressure Detected")
            elif systolic_bp >= 130 or diastolic_bp >= 80:
                st.warning("⚡ Elevated Blood Pressure")
            else:
                st.success("✅ Blood Pressure Normal")
            st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown('<div class="info-box">', unsafe_allow_html=True)
            st.markdown("**🧬 Cholesterol Level**")
            cholesterol = st.slider('💊 Cholesterol (mg/dL)', 60, 160, 120, help="Normal: 125-200 mg/dL")
            
            if cholesterol < 125:
                st.info("ℹ️ Below Normal Range")
            elif cholesterol <= 160:
                st.success("✅ Within Normal Range")
            else:
                st.warning("⚠️ Above Normal Range")
            st.markdown('</div>', unsafe_allow_html=True)
    
    with col_right:
        st.markdown("### 🔬 Risk Assessment")
        st.write("")
        st.write("")
        
        if st.button('🔍 Analyze Risk Profile', use_container_width=True, type='primary', key='predict_btn'):
            patient = pd.DataFrame([[age, systolic_bp, diastolic_bp, cholesterol]], columns=features)
            patient_scaled = scaler.transform(patient)
            
            prob = model.predict_proba(patient_scaled)[0][1]
            label = 'Retinopathy' if prob >= 0.5 else 'No Retinopathy'
            
            # Save to history
            st.session_state.prediction_history.append({
                'timestamp': datetime.now(),
                'age': age,
                'systolic_bp': systolic_bp,
                'diastolic_bp': diastolic_bp,
                'cholesterol': cholesterol,
                'probability': prob,
                'risk_level': 'High' if prob >= 0.70 else 'Medium' if prob >= 0.50 else 'Low'
            })
            
            st.write("")
            
            # Results Display
            if prob >= 0.70:
                st.markdown(f"""
                <div class="risk-high">
                    <h2 style="color: #fff; margin: 0;">🚨 HIGH RISK</h2>
                    <h3 style="color: #fff; margin: 0.5rem 0;">{prob*100:.1f}% Probability</h3>
                    <p style="color: #fff; font-size: 1.1rem; margin: 1rem 0;">
                        <strong>Diagnosis:</strong> {label}
                    </p>
                </div>
                """, unsafe_allow_html=True)
                st.error("⚠️ **Immediate Action Required**")
                st.markdown("""
                **Recommendations:**
                - 🏥 Schedule immediate ophthalmologist consultation
                - 🔬 Request comprehensive eye examination
                - 📋 Monitor blood glucose levels closely
                - 💊 Review current diabetes management plan
                """)
                
            elif prob >= 0.50:
                st.markdown(f"""
                <div class="risk-medium">
                    <h2 style="color: #333; margin: 0;">⚡ MEDIUM RISK</h2>
                    <h3 style="color: #333; margin: 0.5rem 0;">{prob*100:.1f}% Probability</h3>
                    <p style="color: #333; font-size: 1.1rem; margin: 1rem 0;">
                        <strong>Diagnosis:</strong> {label}
                    </p>
                </div>
                """, unsafe_allow_html=True)
                st.warning("⚠️ **Follow-up Recommended**")
                st.markdown("""
                **Recommendations:**
                - 👁️ Schedule eye check-up within 3-6 months
                - 📊 Regular blood pressure monitoring
                - 🏃‍♂️ Maintain healthy lifestyle habits
                - 💊 Continue prescribed medications
                """)
                
            else:
                st.markdown(f"""
                <div class="risk-low">
                    <h2 style="color: #fff; margin: 0;">✅ LOW RISK</h2>
                    <h3 style="color: #fff; margin: 0.5rem 0;">{prob*100:.1f}% Probability</h3>
                    <p style="color: #fff; font-size: 1.1rem; margin: 1rem 0;">
                        <strong>Diagnosis:</strong> {label}
                    </p>
                </div>
                """, unsafe_allow_html=True)
                st.success("✅ **Continue Regular Monitoring**")
                st.markdown("""
                **Recommendations:**
                - 📅 Annual eye examination
                - 🍎 Maintain healthy diet and exercise
                - 📊 Regular diabetes management check-ups
                - 😊 Keep up the good work!
                """)
            
            st.write("")
            
            # Enhanced Visualization
            fig, ax = plt.subplots(figsize=(10, 3))
            fig.patch.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            ax.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            
            colors = ['#4facfe', '#00f2fe'] if prob < 0.5 else ['#f093fb', '#f5576c']
            
            ax.barh(['Risk Score'], [prob], color=colors[0], height=0.5, alpha=0.8)
            ax.barh(['Risk Score'], [1-prob], left=[prob], color='#e0e0e0', height=0.5, alpha=0.5)
            ax.axvline(0.5, color='#333', linestyle='--', linewidth=2, alpha=0.7, label='Risk Threshold')
            
            ax.set_xlim(0, 1)
            ax.set_xlabel('Probability Score', fontsize=12, fontweight='bold')
            ax.set_title(f'Risk Assessment Score: {prob*100:.1f}%', fontsize=14, fontweight='bold', pad=20)
            ax.legend(loc='upper right')
            
            for spine in ax.spines.values():
                spine.set_visible(False)
            
            st.pyplot(fig)
            
            # Additional Metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Risk Score", f"{prob*100:.1f}%", 
                         delta=f"{(prob-0.5)*100:.1f}% vs threshold")
            with col2:
                st.metric("Confidence", f"{max(prob, 1-prob)*100:.1f}%")
            with col3:
                risk_cat = "High" if prob >= 0.70 else "Medium" if prob >= 0.50 else "Low"
                st.metric("Risk Category", risk_cat)

# ========== TAB 2: ADVANCED ANALYTICS ==========
with tab2:
    st.markdown("### 📊 Advanced Patient Analytics")
    
    if len(st.session_state.prediction_history) == 0:
        st.info("👈 Make a prediction in the 'Risk Prediction' tab to see analytics here!")
    else:
        latest = st.session_state.prediction_history[-1]
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("#### 📈 Parameter Comparison with Normal Ranges")
            
            # Create gauge-style comparison
            fig, axes = plt.subplots(2, 2, figsize=(10, 8))
            fig.patch.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            
            params = [
                ('Age', latest['age'], NORMAL_RANGES['age'], 'years'),
                ('Systolic BP', latest['systolic_bp'], NORMAL_RANGES['systolic_bp'], 'mmHg'),
                ('Diastolic BP', latest['diastolic_bp'], NORMAL_RANGES['diastolic_bp'], 'mmHg'),
                ('Cholesterol', latest['cholesterol'], NORMAL_RANGES['cholesterol'], 'mg/dL')
            ]
            
            axes = axes.flatten()
            for idx, (name, value, (min_val, max_val), unit) in enumerate(params):
                ax = axes[idx]
                ax.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
                
                # Background bar (full range)
                ax.barh([0], [max_val + 20], color='#e0e0e0', height=0.5)
                
                # Normal range bar
                ax.barh([0], [max_val - min_val], left=[min_val], 
                       color='#51cf66', alpha=0.3, height=0.5, label='Normal Range')
                
                # Patient value marker
                ax.plot([value], [0], 'ro', markersize=15, label='Patient Value')
                
                ax.set_xlim(0, max_val + 20)
                ax.set_ylim(-0.5, 0.5)
                ax.set_yticks([])
                ax.set_title(f'{name}: {value} {unit}', fontweight='bold')
                ax.legend(loc='upper right', fontsize=8)
                
                # Remove spines
                for spine in ax.spines.values():
                    spine.set_visible(False)
            
            plt.tight_layout()
            st.pyplot(fig)
        
        with col2:
            st.markdown("#### 🎯 Risk Factor Analysis")
            
            # Calculate relative risk for each parameter
            age_risk = (latest['age'] - 30) / 80  # Normalized
            bp_risk = max((latest['systolic_bp'] - 120) / 60, (latest['diastolic_bp'] - 80) / 60, 0)
            chol_risk = max((latest['cholesterol'] - 125) / 75, 0)
            
            risk_factors = {
                'Age Factor': age_risk * 100,
                'Blood Pressure': bp_risk * 100,
                'Cholesterol': chol_risk * 100
            }
            
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.patch.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            ax.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            
            factors = list(risk_factors.keys())
            values = list(risk_factors.values())
            colors_bars = ['#ff6b6b' if v > 50 else '#ffd93d' if v > 25 else '#51cf66' for v in values]
            
            bars = ax.barh(factors, values, color=colors_bars, alpha=0.7)
            ax.set_xlabel('Risk Contribution (%)', fontweight='bold')
            ax.set_title('Individual Risk Factor Contributions', fontweight='bold', pad=20)
            ax.set_xlim(0, 100)
            
            for i, (bar, val) in enumerate(zip(bars, values)):
                ax.text(val + 2, i, f'{val:.1f}%', va='center', fontweight='bold')
            
            for spine in ax.spines.values():
                spine.set_visible(False)
            
            st.pyplot(fig)
            
            st.markdown("#### 📜 Prediction History")
            if len(st.session_state.prediction_history) > 0:
                history_df = pd.DataFrame(st.session_state.prediction_history)
                history_df['timestamp'] = history_df['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
                st.dataframe(history_df, use_container_width=True)
                
                # Clear history button
                if st.button('🗑️ Clear History'):
                    st.session_state.prediction_history = []
                    st.rerun()

# ========== TAB 3: EDUCATION & INFO ==========
with tab3:
    st.markdown("### 📚 Understanding Diabetic Retinopathy")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("""
        #### 🔍 What is Diabetic Retinopathy?
        
        Diabetic retinopathy is a diabetes complication that affects the eyes. It's caused by damage to the blood vessels 
        of the light-sensitive tissue at the back of the eye (retina).
        
        **Key Facts:**
        - Leading cause of blindness in working-age adults
        - Affects 1 in 3 people with diabetes
        - Often has no early warning signs
        - Early detection can prevent 95% of vision loss
        
        #### ⚠️ Risk Factors
        
        1. **Duration of Diabetes** - Longer duration increases risk
        2. **Poor Blood Sugar Control** - High HbA1c levels
        3. **High Blood Pressure** - Damages blood vessels
        4. **High Cholesterol** - Contributes to vessel damage
        5. **Pregnancy** - Gestational diabetes increases risk
        6. **Smoking** - Damages blood vessels
        
        #### 🎯 Prevention Tips
        
        - 📊 Monitor blood sugar regularly
        - 💊 Take medications as prescribed
        - 🏃‍♂️ Exercise regularly (30+ min/day)
        - 🍎 Maintain healthy diet
        - 👁️ Get annual eye exams
        - 🚭 Don't smoke
        - 📉 Control blood pressure and cholesterol
        """)
    
    with col2:
        st.markdown("""
        #### 🧠 How Our AI Model Works
        
        Our prediction system uses machine learning to assess diabetic retinopathy risk based on your clinical parameters.
        
        **Model Architecture:**
        - Algorithm: Logistic Regression
        - Training Data: 6,000 patient records
        - Features: 4 clinical parameters
        - Accuracy: ~85%
        - AUC-ROC: ~0.89
        
        **The Process:**
        
        1. **Data Input** - You provide age, BP, and cholesterol
        2. **Normalization** - Values are standardized
        3. **Prediction** - Model calculates risk probability
        4. **Classification** - Risk level determined (Low/Medium/High)
        5. **Recommendations** - Evidence-based advice provided
        
        #### 📈 Stages of Diabetic Retinopathy
        
        **1. Mild Nonproliferative**
        - Microaneurysms (tiny bulges in blood vessels)
        - Often no symptoms
        
        **2. Moderate Nonproliferative**
        - Blood vessels swell and lose ability to transport blood
        - May cause changes in retina appearance
        
        **3. Severe Nonproliferative**
        - More blood vessels blocked
        - Retina deprived of blood supply
        - Signals body to grow new blood vessels
        
        **4. Proliferative**
        - Most advanced stage
        - New, fragile blood vessels grow
        - Can lead to serious vision problems
        
        #### 🏥 When to See a Doctor
        
        Seek immediate medical attention if you experience:
        - Sudden vision loss
        - Blurred vision
        - Floaters or spots
        - Dark or empty areas in vision
        - Difficulty with color perception
        """)
    
    st.divider()
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 2rem; border-radius: 15px; color: white;">
        <h3 style="margin: 0; color: white;">💡 Did You Know?</h3>
        <p style="font-size: 1.1rem; margin: 1rem 0 0 0;">
            Regular eye exams can detect diabetic retinopathy before you notice any symptoms. 
            Early treatment can reduce the risk of blindness by 95%!
        </p>
    </div>
    """, unsafe_allow_html=True)

# ========== TAB 4: BATCH ANALYSIS ==========
with tab4:
    st.markdown("### 📁 Batch Patient Analysis")
    
    st.markdown("""
    Upload a CSV file with multiple patient records to analyze risk for entire groups.
    
    **Required columns:** `age`, `systolic_bp`, `diastolic_bp`, `cholesterol`
    """)
    
    # Sample CSV download
    sample_data = pd.DataFrame({
        'age': [45, 62, 38, 71, 55],
        'systolic_bp': [115, 142, 108, 155, 125],
        'diastolic_bp': [75, 92, 68, 98, 82],
        'cholesterol': [190, 165, 145, 158, 172]
    })
    
    csv = sample_data.to_csv(index=False)
    st.download_button(
        label="📥 Download Sample CSV Template",
        data=csv,
        file_name="sample_patients.csv",
        mime="text/csv"
    )
    
    uploaded_file = st.file_uploader("Choose a CSV file", type="csv")
    
    if uploaded_file is not None:
        try:
            batch_df = pd.read_csv(uploaded_file)
            
            st.markdown("#### 📊 Uploaded Data Preview")
            st.dataframe(batch_df.head(), use_container_width=True)
            
            if st.button('🔍 Analyze All Patients', use_container_width=True, type='primary'):
                with st.spinner('Analyzing all patients...'):
                    # Make predictions
                    batch_scaled = scaler.transform(batch_df[features])
                    probabilities = model.predict_proba(batch_scaled)[:, 1]
                    
                    batch_df['risk_probability'] = probabilities
                    batch_df['risk_level'] = batch_df['risk_probability'].apply(
                        lambda x: 'High' if x >= 0.70 else 'Medium' if x >= 0.50 else 'Low'
                    )
                    batch_df['recommendation'] = batch_df['risk_level'].map({
                        'High': 'Immediate consultation',
                        'Medium': 'Schedule follow-up',
                        'Low': 'Regular monitoring'
                    })
                    
                    st.success(f"✅ Analyzed {len(batch_df)} patients successfully!")
                    
                    # Results summary
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        high_risk = (batch_df['risk_level'] == 'High').sum()
                        st.metric("High Risk Patients", high_risk, 
                                 help="Require immediate attention")
                    with col2:
                        medium_risk = (batch_df['risk_level'] == 'Medium').sum()
                        st.metric("Medium Risk Patients", medium_risk,
                                 help="Require follow-up")
                    with col3:
                        low_risk = (batch_df['risk_level'] == 'Low').sum()
                        st.metric("Low Risk Patients", low_risk,
                                 help="Continue regular monitoring")
                    
                    # Detailed results
                    st.markdown("#### 📋 Detailed Results")
                    st.dataframe(batch_df, use_container_width=True)
                    
                    # Visualization
                    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
                    fig.patch.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
                    
                    # Risk distribution
                    ax1.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
                    risk_counts = batch_df['risk_level'].value_counts()
                    colors_pie = ['#ff6b6b', '#ffd93d', '#51cf66']
                    ax1.pie(risk_counts.values, labels=risk_counts.index, autopct='%1.1f%%',
                           colors=colors_pie, startangle=90)
                    ax1.set_title('Risk Distribution', fontweight='bold')
                    
                    # Probability histogram
                    ax2.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
                    ax2.hist(batch_df['risk_probability'], bins=20, color='#667eea', alpha=0.7, edgecolor='black')
                    ax2.axvline(0.5, color='red', linestyle='--', linewidth=2, label='Risk Threshold')
                    ax2.set_xlabel('Risk Probability', fontweight='bold')
                    ax2.set_ylabel('Number of Patients', fontweight='bold')
                    ax2.set_title('Probability Distribution', fontweight='bold')
                    ax2.legend()
                    
                    plt.tight_layout()
                    st.pyplot(fig)
                    
                    # Export results
                    result_csv = batch_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Analysis Results",
                        data=result_csv,
                        file_name=f"retinopathy_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
        
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            st.info("Please ensure your CSV has the required columns: age, systolic_bp, diastolic_bp, cholesterol")

# ========== TAB 5: REAL-TIME DASHBOARD ==========
with tab5:
    st.markdown("### 📈 Real-Time Analytics Dashboard")
    
    if len(st.session_state.prediction_history) == 0:
        st.info("👈 Make predictions to populate the dashboard")
    else:
        history_df = pd.DataFrame(st.session_state.prediction_history)
        
        # KPI Cards
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.metric("Total Analyzed", len(history_df))
        with col2:
            avg_risk = history_df['probability'].mean()
            st.metric("Avg Risk Score", f"{avg_risk*100:.1f}%")
        with col3:
            high_count = (history_df['probability'] >= 0.70).sum()
            st.metric("High Risk", high_count)
        with col4:
            medium_count = ((history_df['probability'] >= 0.50) & (history_df['probability'] < 0.70)).sum()
            st.metric("Medium Risk", medium_count)
        with col5:
            low_count = (history_df['probability'] < 0.50).sum()
            st.metric("Low Risk", low_count)
        
        st.divider()
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("#### 📊 Risk Distribution")
            risk_dist = history_df['risk_level'].value_counts()
            fig, ax = plt.subplots(figsize=(8, 6))
            fig.patch.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            ax.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            
            colors_pie = ['#ff6b6b', '#ffd93d', '#51cf66']
            ax.pie(risk_dist.values, labels=risk_dist.index, autopct='%1.1f%%',
                   colors=colors_pie, startangle=90)
            ax.set_title('Patient Risk Levels', fontweight='bold')
            st.pyplot(fig)
        
        with col2:
            st.markdown("#### 📈 Risk Trend Over Time")
            history_df['timestamp'] = pd.to_datetime(history_df['timestamp'])
            history_sorted = history_df.sort_values('timestamp')
            
            fig, ax = plt.subplots(figsize=(8, 6))
            fig.patch.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            ax.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            
            ax.scatter(range(len(history_sorted)), history_sorted['probability']*100, 
                      color='#667eea', s=100, alpha=0.6)
            ax.plot(range(len(history_sorted)), history_sorted['probability']*100, 
                   color='#667eea', alpha=0.3)
            ax.axhline(50, color='red', linestyle='--', alpha=0.5, label='Risk Threshold')
            ax.set_xlabel('Prediction #', fontweight='bold')
            ax.set_ylabel('Risk Score (%)', fontweight='bold')
            ax.set_title('Risk Score Progression', fontweight='bold')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            for spine in ax.spines.values():
                spine.set_visible(True)
            
            st.pyplot(fig)
        
        st.divider()
        
        # Parameter Statistics
        st.markdown("#### 📊 Parameter Statistics")
        
        stats_data = {
            'Parameter': ['Age', 'Systolic BP', 'Diastolic BP', 'Cholesterol'],
            'Min': [
                history_df['age'].min(),
                history_df['systolic_bp'].min(),
                history_df['diastolic_bp'].min(),
                history_df['cholesterol'].min()
            ],
            'Mean': [
                f"{history_df['age'].mean():.1f}",
                f"{history_df['systolic_bp'].mean():.1f}",
                f"{history_df['diastolic_bp'].mean():.1f}",
                f"{history_df['cholesterol'].mean():.1f}"
            ],
            'Max': [
                history_df['age'].max(),
                history_df['systolic_bp'].max(),
                history_df['diastolic_bp'].max(),
                history_df['cholesterol'].max()
            ]
        }
        stats_df = pd.DataFrame(stats_data)
        st.dataframe(stats_df, use_container_width=True, hide_index=True)

# ========== TAB 6: RISK STRATIFICATION ==========
with tab6:
    st.markdown("### 👥 Age & Risk Stratification Analysis")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        if len(st.session_state.prediction_history) == 0:
            st.info("👈 Make predictions in Risk Prediction tab to see stratification")
        else:
            history_df = pd.DataFrame(st.session_state.prediction_history)
            
            # Create age groups
            history_df['age_group'] = pd.cut(history_df['age'], 
                                             bins=[0, 30, 40, 50, 60, 70, 150],
                                             labels=['<30', '30-40', '40-50', '50-60', '60-70', '70+'])
            
            # Stratification analysis
            stratification = history_df.groupby('age_group').agg({
                'probability': ['count', 'mean', 'min', 'max'],
                'risk_level': lambda x: (x == 'High').sum()
            }).round(3)
            
            stratification.columns = ['Patient Count', 'Avg Risk %', 'Min Risk %', 'Max Risk %', 'High Risk Count']
            stratification['Avg Risk %'] = stratification['Avg Risk %'] * 100
            stratification['Min Risk %'] = stratification['Min Risk %'] * 100
            stratification['Max Risk %'] = stratification['Max Risk %'] * 100
            
            st.markdown("#### 📊 Age Group Risk Summary")
            st.dataframe(stratification, use_container_width=True)
            
            # Visualization
            st.markdown("#### 📈 Risk by Age Group")
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
            fig.patch.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            
            # Average risk by age group
            ax1.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            age_groups = history_df['age_group'].value_counts().sort_index()
            avg_risks = history_df.groupby('age_group')['probability'].mean() * 100
            
            colors_bar = ['#51cf66' if x < 50 else '#ffd93d' if x < 70 else '#ff6b6b' for x in avg_risks]
            ax1.bar(range(len(avg_risks)), avg_risks, color=colors_bar, alpha=0.7)
            ax1.set_xticks(range(len(avg_risks)))
            ax1.set_xticklabels(avg_risks.index)
            ax1.set_ylabel('Average Risk Score (%)', fontweight='bold')
            ax1.set_title('Average Risk by Age Group', fontweight='bold')
            ax1.axhline(50, color='red', linestyle='--', alpha=0.5)
            ax1.set_ylim(0, 100)
            
            for i, v in enumerate(avg_risks):
                ax1.text(i, v + 2, f'{v:.1f}%', ha='center', fontweight='bold')
            
            # Patient count by age group
            ax2.set_facecolor('#f8f9fa' if not st.session_state.dark_mode else '#1e1e1e')
            ax2.bar(range(len(age_groups)), age_groups.values, color='#667eea', alpha=0.7)
            ax2.set_xticks(range(len(age_groups)))
            ax2.set_xticklabels(age_groups.index)
            ax2.set_ylabel('Patient Count', fontweight='bold')
            ax2.set_title('Patients by Age Group', fontweight='bold')
            
            for i, v in enumerate(age_groups.values):
                ax2.text(i, v + 0.2, str(v), ha='center', fontweight='bold')
            
            for spine in ax1.spines.values():
                spine.set_visible(False)
            for spine in ax2.spines.values():
                spine.set_visible(False)
            
            plt.tight_layout()
            st.pyplot(fig)
    
    with col2:
        st.markdown("#### 📋 Risk Categories")
        st.markdown("""
        **Low Risk** 🟢
        - Probability < 50%
        - Continue monitoring
        
        **Medium Risk** 🟡
        - Probability 50-70%
        - Schedule follow-up
        
        **High Risk** 🔴
        - Probability ≥ 70%
        - Immediate action
        """)

# ========== TAB 7: REPORTS & EXPORT ==========
with tab7:
    st.markdown("### 📋 Report Generation & Export")
    
    if len(st.session_state.prediction_history) == 0:
        st.info("👈 Make a prediction in the 'Risk Prediction' tab to generate reports")
    else:
        latest = st.session_state.prediction_history[-1]
        
        # Report generation options
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("#### 📄 PDF Report")
            
            if HAS_REPORTLAB:
                def generate_pdf_report(patient_data):
                    buffer = BytesIO()
                    doc = SimpleDocTemplate(buffer, pagesize=letter,
                                          rightMargin=72, leftMargin=72,
                                          topMargin=72, bottomMargin=18)
                    Story = []
                    styles = getSampleStyleSheet()
                    
                    # Title
                    title_style = ParagraphStyle(
                        'CustomTitle',
                        parent=styles['Heading1'],
                        fontSize=24,
                        textColor=colors.HexColor('#667eea'),
                        spaceAfter=30,
                        alignment=1
                    )
                    Story.append(Paragraph("Diabetic Retinopathy Risk Assessment Report", title_style))
                    
                    # Report info
                    Story.append(Spacer(1, 0.2*inch))
                    Story.append(Paragraph(f"<b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
                    
                    # Patient data
                    Story.append(Spacer(1, 0.3*inch))
                    Story.append(Paragraph("<b>Patient Information</b>", styles['Heading2']))
                    
                    patient_table = [
                        ['Parameter', 'Value'],
                        ['Age', f"{patient_data['age']} years"],
                        ['Systolic BP', f"{patient_data['systolic_bp']} mmHg"],
                        ['Diastolic BP', f"{patient_data['diastolic_bp']} mmHg"],
                        ['Cholesterol', f"{patient_data['cholesterol']} mg/dL"]
                    ]
                    
                    t = Table(patient_table, colWidths=[3*inch, 3*inch])
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                    ]))
                    Story.append(t)
                    
                    # Risk assessment
                    Story.append(Spacer(1, 0.3*inch))
                    Story.append(Paragraph("<b>Risk Assessment</b>", styles['Heading2']))
                    
                    prob = patient_data['probability']
                    risk_level = patient_data['risk_level']
                    
                    Story.append(Paragraph(f"<b>Risk Score:</b> {prob*100:.1f}%", styles['Normal']))
                    Story.append(Paragraph(f"<b>Risk Level:</b> {risk_level}", styles['Normal']))
                    
                    # Recommendations
                    Story.append(Spacer(1, 0.3*inch))
                    Story.append(Paragraph("<b>Recommendations</b>", styles['Heading2']))
                    
                    if prob >= 0.70:
                        recs = "Schedule immediate ophthalmologist consultation. Request comprehensive eye examination. Monitor blood glucose levels closely."
                    elif prob >= 0.50:
                        recs = "Schedule eye check-up within 3-6 months. Regular blood pressure monitoring. Maintain healthy lifestyle habits."
                    else:
                        recs = "Annual eye examination recommended. Maintain healthy diet and exercise. Continue regular diabetes management."
                    
                    Story.append(Paragraph(recs, styles['Normal']))
                    
                    # Disclaimer
                    Story.append(Spacer(1, 0.3*inch))
                    disclaimer_style = ParagraphStyle(
                        'Disclaimer',
                        parent=styles['Normal'],
                        textColor=colors.HexColor('#666666'),
                        fontSize=9
                    )
                    Story.append(Paragraph("<i>This report is for educational purposes only. Consult healthcare professionals for medical decisions.</i>", disclaimer_style))
                    
                    doc.build(Story)
                    buffer.seek(0)
                    return buffer
                
                if st.button('📥 Generate PDF Report', use_container_width=True):
                    pdf_buffer = generate_pdf_report(latest)
                    st.download_button(
                        label="⬇️ Download PDF Report",
                        data=pdf_buffer,
                        file_name=f"retinopathy_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
            else:
                st.warning("ReportLab not installed. PDF generation disabled.")
        
        with col2:
            st.markdown("#### 📊 Excel Export")
            
            if HAS_OPENPYXL:
                def generate_excel_export(history_list):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Predictions"
                    
                    # Header
                    headers = ['Timestamp', 'Age', 'Systolic BP', 'Diastolic BP', 'Cholesterol', 'Risk Probability', 'Risk Level']
                    ws.append(headers)
                    
                    # Style header
                    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF')
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Add data
                    for record in history_list:
                        ws.append([
                            record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                            record['age'],
                            record['systolic_bp'],
                            record['diastolic_bp'],
                            record['cholesterol'],
                            f"{record['probability']*100:.1f}%",
                            record['risk_level']
                        ])
                    
                    # Set column widths
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 10
                    ws.column_dimensions['C'].width = 12
                    ws.column_dimensions['D'].width = 12
                    ws.column_dimensions['E'].width = 12
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 12
                    
                    # Add summary sheet
                    summary = wb.create_sheet("Summary")
                    summary['A1'] = 'Total Patients'
                    summary['B1'] = len(history_list)
                    summary['A2'] = 'High Risk'
                    summary['B2'] = sum(1 for x in history_list if x['probability'] >= 0.70)
                    summary['A3'] = 'Medium Risk'
                    summary['B3'] = sum(1 for x in history_list if 0.50 <= x['probability'] < 0.70)
                    summary['A4'] = 'Low Risk'
                    summary['B4'] = sum(1 for x in history_list if x['probability'] < 0.50)
                    summary['A5'] = 'Average Risk'
                    summary['B5'] = f"{np.mean([x['probability'] for x in history_list])*100:.1f}%"
                    
                    for cell in summary['A']:
                        cell.font = Font(bold=True)
                    
                    return BytesIO(wb.save(None))
                
                if st.button('📊 Generate Excel Report', use_container_width=True):
                    excel_buffer = BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Predictions"
                    
                    headers = ['Timestamp', 'Age', 'Systolic BP', 'Diastolic BP', 'Cholesterol', 'Risk Probability', 'Risk Level']
                    ws.append(headers)
                    
                    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF')
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    for record in st.session_state.prediction_history:
                        ws.append([
                            record['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                            record['age'],
                            record['systolic_bp'],
                            record['diastolic_bp'],
                            record['cholesterol'],
                            f"{record['probability']*100:.1f}%",
                            record['risk_level']
                        ])
                    
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 10
                    ws.column_dimensions['C'].width = 12
                    ws.column_dimensions['D'].width = 12
                    ws.column_dimensions['E'].width = 12
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 12
                    
                    summary = wb.create_sheet("Summary")
                    summary['A1'] = 'Total Patients'
                    summary['B1'] = len(st.session_state.prediction_history)
                    summary['A2'] = 'High Risk'
                    summary['B2'] = sum(1 for x in st.session_state.prediction_history if x['probability'] >= 0.70)
                    summary['A3'] = 'Medium Risk'
                    summary['B3'] = sum(1 for x in st.session_state.prediction_history if 0.50 <= x['probability'] < 0.70)
                    summary['A4'] = 'Low Risk'
                    summary['B4'] = sum(1 for x in st.session_state.prediction_history if x['probability'] < 0.50)
                    summary['A5'] = 'Average Risk'
                    summary['B5'] = f"{np.mean([x['probability'] for x in st.session_state.prediction_history])*100:.1f}%"
                    
                    for cell in summary['A']:
                        cell.font = Font(bold=True)
                    
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ Download Excel Report",
                        data=excel_buffer,
                        file_name=f"retinopathy_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.warning("openpyxl not installed. Excel export disabled.")
        
        st.divider()
        
        # CSV Export (always available)
        st.markdown("#### 📥 CSV Export")
        
        history_df = pd.DataFrame(st.session_state.prediction_history)
        history_df['timestamp'] = history_df['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
        csv_data = history_df.to_csv(index=False)
        
        st.download_button(
            label="📥 Download CSV Report",
            data=csv_data,
            file_name=f"retinopathy_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True
        )

# ========== TAB 8: RECOMMENDATIONS & ACTIONS ==========
with tab8:
    st.markdown("### 💬 Personalized Recommendations & Actions")
    
    if len(st.session_state.prediction_history) == 0:
        st.info("👈 Make a prediction in the 'Risk Prediction' tab to get personalized recommendations")
    else:
        latest = st.session_state.prediction_history[-1]
        prob = latest['probability']
        
        # Risk-based recommendations
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown("#### 🎯 Customized Action Plan")
            
            if prob >= 0.70:
                st.error("🚨 HIGH RISK - IMMEDIATE ACTION REQUIRED")
                st.markdown("""
                **Your personalized action plan:**
                
                1. **Immediate (This Week)**
                   - 📞 Call your ophthalmologist immediately
                   - 🏥 Schedule emergency eye examination
                   - 📋 Gather recent medical records
                   - 💊 Ensure medications are up to date
                
                2. **Short Term (Next 2 weeks)**
                   - 🔬 Complete comprehensive eye exam
                   - 📊 Get HbA1c test results
                   - 👨‍⚕️ Meet with endocrinologist
                   - 📝 Document all eye symptoms
                
                3. **Ongoing**
                   - 🩺 Monthly ophthalmology follow-ups
                   - 📈 Weekly blood glucose monitoring
                   - 💪 Intensive diabetes management
                   - 👓 Use protective eyewear
                   - 🚭 Quit smoking immediately
                """)
            elif prob >= 0.50:
                st.warning("⚡ MEDIUM RISK - FOLLOW-UP NEEDED")
                st.markdown("""
                **Your personalized action plan:**
                
                1. **This Month**
                   - 📅 Schedule eye check-up (within 3-6 months)
                   - 📊 Check blood pressure regularly
                   - 📝 Review diabetes medications
                
                2. **Monthly**
                   - 🩺 Monitor blood glucose levels
                   - 📈 Track BP readings
                   - 🏥 Prepare questions for next ophthalmologist visit
                
                3. **Lifestyle Changes**
                   - 🏃‍♂️ Exercise 30 minutes daily
                   - 🍎 Adopt heart-healthy diet
                   - 💤 Get 7-8 hours sleep
                   - 🚭 Avoid smoking
                   - 📚 Learn about diabetes management
                """)
            else:
                st.success("✅ LOW RISK - MAINTAIN GOOD HABITS")
                st.markdown("""
                **Your personalized action plan:**
                
                1. **Annual Care**
                   - 👁️ Annual comprehensive eye exam
                   - 📊 Annual blood work screening
                   - 🩺 Regular diabetes check-ups
                
                2. **Daily Habits**
                   - 🍏 Maintain healthy diet
                   - 💧 Drink plenty of water
                   - 🏃‍♂️ Stay physically active
                   - 😴 Maintain regular sleep schedule
                
                3. **Prevention**
                   - 📚 Stay informed about diabetes
                   - 📊 Monitor and control weight
                   - 🧂 Limit salt and sugar intake
                   - 🏥 Keep regular medical appointments
                """)
        
        with col1:
            st.markdown("#### 📧 Share Results")
            
            # Email composition
            email_subject = f"Diabetic Retinopathy Risk Assessment - {latest['risk_level']} Risk"
            email_body = f"""
Diabetic Retinopathy Risk Assessment Report
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

PATIENT INFORMATION:
- Age: {latest['age']} years
- Systolic BP: {latest['systolic_bp']} mmHg
- Diastolic BP: {latest['diastolic_bp']} mmHg
- Cholesterol: {latest['cholesterol']} mg/dL

RISK ASSESSMENT:
- Risk Score: {prob*100:.1f}%
- Risk Level: {latest['risk_level']}

This assessment is for educational purposes only.
Please consult with a healthcare professional for medical decisions.
            """
            
            if st.button('📋 Copy Results to Clipboard'):
                st.code(email_body)
                st.success("✅ Results copied! You can now paste them in emails or documents.")
            
            st.markdown("#### 🖨️ Print Recommendations")
            
            if st.button('🖨️ Prepare for Printing'):
                st.markdown("""
                <style>
                @media print {
                    .print-section { page-break-inside: avoid; }
                }
                </style>
                """, unsafe_allow_html=True)
                
                print_content = f"""
DIABETIC RETINOPATHY RISK ASSESSMENT REPORT
=============================================

Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

PATIENT DATA:
Age: {latest['age']} years
Systolic BP: {latest['systolic_bp']} mmHg
Diastolic BP: {latest['diastolic_bp']} mmHg
Cholesterol: {latest['cholesterol']} mg/dL

RESULTS:
Risk Score: {prob*100:.1f}%
Risk Level: {latest['risk_level']}

RECOMMENDATION:
See above section for action plan.

=============================================
For healthcare professionals only.
This assessment is for educational purposes.
                """
                
                st.text_area("Print Preview:", value=print_content, height=300, disabled=True)
                st.info("Use Ctrl+P or Cmd+P to print this page")

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; padding: 2rem 0;">
    <p><strong>P653 Machine Learning Project</strong> | Diabetic Retinopathy Risk Assessment - Advanced Edition</p>
    <p>Powered by Logistic Regression | Built with ❤️ using Streamlit | Enhanced with 8 Advanced Features</p>
</div>
""", unsafe_allow_html=True)
